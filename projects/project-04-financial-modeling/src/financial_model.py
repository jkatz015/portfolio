"""
Financial Model - 3-Statement Model with Scenario Analysis
==========================================================
Builds Income Statement, Balance Sheet, and Cash Flow projections
for a restaurant business with Base, Upside, and Downside scenarios.

Usage:
    python financial_model.py

Output:
    - model/financial_model.xlsx (multi-tab Excel workbook)
"""

import pandas as pd
import numpy as np
from pathlib import Path

# ============================================================
# CONFIGURATION
# ============================================================

# File paths
BASE_DIR = Path(__file__).parent.parent
DATA_FILE = BASE_DIR / "data" / "historical_financials.csv"
OUTPUT_FILE = BASE_DIR / "model" / "financial_model.xlsx"

# Projection period
PROJECTION_YEARS = [2024, 2025, 2026, 2027, 2028]
HISTORICAL_YEARS = [2022, 2023]

# Starting balance sheet (end of 2023)
STARTING_BALANCE = {
    'cash': 85000,
    'accounts_receivable': 18000,
    'inventory': 42000,
    'fixed_assets_net': 180000,
    'accounts_payable': 35000,
    'current_debt': 30000,
    'long_term_debt': 90000,
    'retained_earnings': 170000,
}

# Working capital assumptions (days)
WORKING_CAPITAL = {
    'ar_days': 3,
    'inventory_days': 7,
    'ap_days': 14,
}

# Depreciation (existing assets)
EXISTING_DEPRECIATION = 25000  # Annual depreciation on existing assets

# Debt assumptions
DEBT_INTEREST_RATE = 0.07
MONTHLY_DEBT_PAYMENT = 2500

# Tax rate
TAX_RATE = 0.25

# ============================================================
# SCENARIO ASSUMPTIONS
# ============================================================

SCENARIOS = {
    'Base': {
        'revenue_growth': [0.06, 0.05, 0.045, 0.04, 0.035],
        'food_cost_pct': [0.315, 0.31, 0.31, 0.305, 0.305],
        'beverage_cost_pct': [0.09, 0.09, 0.085, 0.085, 0.085],
        'labor_cost_pct': [0.28, 0.275, 0.275, 0.27, 0.27],
        'rent_base': 156000,
        'rent_escalation': 0.03,
        'utilities_pct': 0.02,
        'marketing_pct': 0.02,
        'insurance_monthly': 2000,
        'supplies_pct': 0.015,
        'repairs_pct': 0.01,
        'cc_fees_pct': 0.03,
        'other_opex_pct': 0.015,
        'capex': [25000, 20000, 15000, 30000, 15000],
    },
    'Upside': {
        'revenue_growth': [0.10, 0.12, 0.10, 0.08, 0.07],
        'food_cost_pct': [0.31, 0.305, 0.30, 0.295, 0.29],
        'beverage_cost_pct': [0.085, 0.085, 0.08, 0.08, 0.08],
        'labor_cost_pct': [0.275, 0.27, 0.265, 0.26, 0.26],
        'rent_base': 156000,
        'rent_escalation': 0.03,
        'utilities_pct': 0.019,
        'marketing_pct': 0.025,
        'insurance_monthly': 2000,
        'supplies_pct': 0.014,
        'repairs_pct': 0.009,
        'cc_fees_pct': 0.03,
        'other_opex_pct': 0.014,
        'capex': [35000, 50000, 40000, 100000, 50000],
    },
    'Downside': {
        'revenue_growth': [0.02, 0.00, 0.01, 0.02, 0.03],
        'food_cost_pct': [0.33, 0.34, 0.335, 0.33, 0.325],
        'beverage_cost_pct': [0.095, 0.10, 0.095, 0.095, 0.09],
        'labor_cost_pct': [0.29, 0.30, 0.295, 0.29, 0.285],
        'rent_base': 156000,
        'rent_escalation': 0.03,
        'utilities_pct': 0.022,
        'marketing_pct': 0.015,
        'insurance_monthly': 2000,
        'supplies_pct': 0.016,
        'repairs_pct': 0.012,
        'cc_fees_pct': 0.03,
        'other_opex_pct': 0.016,
        'capex': [15000, 10000, 10000, 15000, 10000],
    },
}


# ============================================================
# LOAD HISTORICAL DATA
# ============================================================

def load_historical_data():
    """Load and summarize historical financial data."""
    df = pd.read_csv(DATA_FILE)

    # Aggregate by year
    yearly = df.groupby('Year').agg({
        'Revenue': 'sum',
        'Food_Cost': 'sum',
        'Beverage_Cost': 'sum',
        'Labor_Cost': 'sum',
        'Rent': 'sum',
        'Utilities': 'sum',
        'Marketing': 'sum',
        'Insurance': 'sum',
        'Supplies': 'sum',
        'Repairs_Maintenance': 'sum',
        'Credit_Card_Fees': 'sum',
        'Other_OpEx': 'sum',
        'CapEx': 'sum',
        'Debt_Payment': 'sum',
    }).reset_index()

    return yearly


def get_base_revenue():
    """Get 2023 revenue as base for projections."""
    df = pd.read_csv(DATA_FILE)
    return df[df['Year'] == 2023]['Revenue'].sum()


# ============================================================
# BUILD INCOME STATEMENT
# ============================================================

def build_income_statement(scenario_name, assumptions, base_revenue):
    """Build 5-year income statement projection."""

    years = PROJECTION_YEARS
    data = {
        'Year': years,
        'Scenario': [scenario_name] * len(years),
    }

    revenue = base_revenue
    revenues = []

    for i, year in enumerate(years):
        # Revenue
        revenue = revenue * (1 + assumptions['revenue_growth'][i])
        revenues.append(revenue)

    data['Revenue'] = revenues

    # Cost of Goods Sold
    data['Food_Cost'] = [r * assumptions['food_cost_pct'][i] for i, r in enumerate(revenues)]
    data['Beverage_Cost'] = [r * assumptions['beverage_cost_pct'][i] for i, r in enumerate(revenues)]
    data['COGS'] = [data['Food_Cost'][i] + data['Beverage_Cost'][i] for i in range(len(years))]

    # Gross Profit
    data['Gross_Profit'] = [data['Revenue'][i] - data['COGS'][i] for i in range(len(years))]
    data['Gross_Margin_Pct'] = [data['Gross_Profit'][i] / data['Revenue'][i] for i in range(len(years))]

    # Operating Expenses
    data['Labor_Cost'] = [r * assumptions['labor_cost_pct'][i] for i, r in enumerate(revenues)]

    # Rent with escalation
    rent = assumptions['rent_base']
    rents = []
    for i in range(len(years)):
        if i > 0:
            rent = rent * (1 + assumptions['rent_escalation'])
        rents.append(rent)
    data['Rent'] = rents

    data['Utilities'] = [r * assumptions['utilities_pct'] for r in revenues]
    data['Marketing'] = [r * assumptions['marketing_pct'] for r in revenues]
    data['Insurance'] = [assumptions['insurance_monthly'] * 12] * len(years)
    data['Supplies'] = [r * assumptions['supplies_pct'] for r in revenues]
    data['Repairs_Maintenance'] = [r * assumptions['repairs_pct'] for r in revenues]
    data['Credit_Card_Fees'] = [r * assumptions['cc_fees_pct'] for r in revenues]
    data['Other_OpEx'] = [r * assumptions['other_opex_pct'] for r in revenues]

    # Total Operating Expenses
    data['Total_OpEx'] = [
        data['Labor_Cost'][i] + data['Rent'][i] + data['Utilities'][i] +
        data['Marketing'][i] + data['Insurance'][i] + data['Supplies'][i] +
        data['Repairs_Maintenance'][i] + data['Credit_Card_Fees'][i] + data['Other_OpEx'][i]
        for i in range(len(years))
    ]

    # EBITDA
    data['EBITDA'] = [data['Gross_Profit'][i] - data['Total_OpEx'][i] for i in range(len(years))]
    data['EBITDA_Margin_Pct'] = [data['EBITDA'][i] / data['Revenue'][i] for i in range(len(years))]

    # Depreciation (existing + new CapEx depreciated over 7 years)
    cumulative_capex = 0
    depreciation = []
    for i in range(len(years)):
        cumulative_capex += assumptions['capex'][i]
        new_depreciation = cumulative_capex / 7  # Simplified: straight-line over 7 years
        depreciation.append(EXISTING_DEPRECIATION + new_depreciation)
    data['Depreciation'] = depreciation

    # EBIT (Operating Income)
    data['EBIT'] = [data['EBITDA'][i] - data['Depreciation'][i] for i in range(len(years))]

    # Interest Expense (simplified - on remaining debt)
    remaining_debt = STARTING_BALANCE['current_debt'] + STARTING_BALANCE['long_term_debt']
    interest = []
    for i in range(len(years)):
        interest.append(remaining_debt * DEBT_INTEREST_RATE)
        remaining_debt = max(0, remaining_debt - (MONTHLY_DEBT_PAYMENT * 12 - remaining_debt * DEBT_INTEREST_RATE))
    data['Interest_Expense'] = interest

    # EBT (Earnings Before Tax)
    data['EBT'] = [data['EBIT'][i] - data['Interest_Expense'][i] for i in range(len(years))]

    # Taxes
    data['Income_Tax'] = [max(0, data['EBT'][i] * TAX_RATE) for i in range(len(years))]

    # Net Income
    data['Net_Income'] = [data['EBT'][i] - data['Income_Tax'][i] for i in range(len(years))]
    data['Net_Margin_Pct'] = [data['Net_Income'][i] / data['Revenue'][i] for i in range(len(years))]

    return pd.DataFrame(data)


# ============================================================
# BUILD BALANCE SHEET
# ============================================================

def build_balance_sheet(scenario_name, income_stmt, assumptions):
    """Build 5-year balance sheet projection."""

    years = PROJECTION_YEARS
    data = {
        'Year': years,
        'Scenario': [scenario_name] * len(years),
    }

    # Initialize from starting balance
    cash = STARTING_BALANCE['cash']
    ar = STARTING_BALANCE['accounts_receivable']
    inventory = STARTING_BALANCE['inventory']
    fixed_assets = STARTING_BALANCE['fixed_assets_net']
    ap = STARTING_BALANCE['accounts_payable']
    current_debt = STARTING_BALANCE['current_debt']
    long_term_debt = STARTING_BALANCE['long_term_debt']
    retained_earnings = STARTING_BALANCE['retained_earnings']

    # Lists to store projections
    cash_list, ar_list, inv_list, fa_list = [], [], [], []
    ap_list, cd_list, ltd_list, re_list = [], [], [], []

    for i, year in enumerate(years):
        revenue = income_stmt.loc[i, 'Revenue']
        cogs = income_stmt.loc[i, 'COGS']
        net_income = income_stmt.loc[i, 'Net_Income']
        depreciation = income_stmt.loc[i, 'Depreciation']
        capex = assumptions['capex'][i]

        # Working capital based on days
        daily_revenue = revenue / 365
        daily_cogs = cogs / 365

        ar = daily_revenue * WORKING_CAPITAL['ar_days']
        inventory = daily_cogs * WORKING_CAPITAL['inventory_days']
        ap = daily_cogs * WORKING_CAPITAL['ap_days']

        # Fixed assets
        fixed_assets = fixed_assets - depreciation + capex

        # Debt paydown
        annual_payment = MONTHLY_DEBT_PAYMENT * 12
        total_debt = current_debt + long_term_debt
        if total_debt > 0:
            interest = total_debt * DEBT_INTEREST_RATE
            principal_payment = min(annual_payment - interest, total_debt)
            total_debt = max(0, total_debt - principal_payment)
            current_debt = min(30000, total_debt)  # Current portion
            long_term_debt = max(0, total_debt - current_debt)

        # Retained earnings
        retained_earnings = retained_earnings + net_income

        # Cash (plug to balance)
        total_assets_ex_cash = ar + inventory + fixed_assets
        total_liabilities = ap + current_debt + long_term_debt
        equity = retained_earnings
        cash = total_liabilities + equity - total_assets_ex_cash
        cash = max(0, cash)  # Can't be negative

        # Store values
        cash_list.append(cash)
        ar_list.append(ar)
        inv_list.append(inventory)
        fa_list.append(fixed_assets)
        ap_list.append(ap)
        cd_list.append(current_debt)
        ltd_list.append(long_term_debt)
        re_list.append(retained_earnings)

    # Assets
    data['Cash'] = cash_list
    data['Accounts_Receivable'] = ar_list
    data['Inventory'] = inv_list
    data['Total_Current_Assets'] = [cash_list[i] + ar_list[i] + inv_list[i] for i in range(len(years))]
    data['Fixed_Assets_Net'] = fa_list
    data['Total_Assets'] = [data['Total_Current_Assets'][i] + fa_list[i] for i in range(len(years))]

    # Liabilities
    data['Accounts_Payable'] = ap_list
    data['Current_Debt'] = cd_list
    data['Total_Current_Liabilities'] = [ap_list[i] + cd_list[i] for i in range(len(years))]
    data['Long_Term_Debt'] = ltd_list
    data['Total_Liabilities'] = [data['Total_Current_Liabilities'][i] + ltd_list[i] for i in range(len(years))]

    # Equity
    data['Retained_Earnings'] = re_list
    data['Total_Equity'] = re_list  # Simplified - no other equity components

    # Check
    data['Total_Liab_Equity'] = [data['Total_Liabilities'][i] + data['Total_Equity'][i] for i in range(len(years))]

    return pd.DataFrame(data)


# ============================================================
# BUILD CASH FLOW STATEMENT
# ============================================================

def build_cash_flow(scenario_name, income_stmt, balance_sheet, assumptions):
    """Build 5-year cash flow statement."""

    years = PROJECTION_YEARS
    data = {
        'Year': years,
        'Scenario': [scenario_name] * len(years),
    }

    # Operating Cash Flow
    net_income = income_stmt['Net_Income'].tolist()
    depreciation = income_stmt['Depreciation'].tolist()

    # Changes in working capital
    ar_change, inv_change, ap_change = [], [], []

    prev_ar = STARTING_BALANCE['accounts_receivable']
    prev_inv = STARTING_BALANCE['inventory']
    prev_ap = STARTING_BALANCE['accounts_payable']

    for i in range(len(years)):
        ar_change.append(-(balance_sheet.loc[i, 'Accounts_Receivable'] - prev_ar))
        inv_change.append(-(balance_sheet.loc[i, 'Inventory'] - prev_inv))
        ap_change.append(balance_sheet.loc[i, 'Accounts_Payable'] - prev_ap)

        prev_ar = balance_sheet.loc[i, 'Accounts_Receivable']
        prev_inv = balance_sheet.loc[i, 'Inventory']
        prev_ap = balance_sheet.loc[i, 'Accounts_Payable']

    data['Net_Income'] = net_income
    data['Depreciation'] = depreciation
    data['Change_in_AR'] = ar_change
    data['Change_in_Inventory'] = inv_change
    data['Change_in_AP'] = ap_change
    data['Operating_Cash_Flow'] = [
        net_income[i] + depreciation[i] + ar_change[i] + inv_change[i] + ap_change[i]
        for i in range(len(years))
    ]

    # Investing Cash Flow
    data['CapEx'] = [-assumptions['capex'][i] for i in range(len(years))]
    data['Investing_Cash_Flow'] = data['CapEx']

    # Financing Cash Flow
    debt_payments = []
    prev_debt = STARTING_BALANCE['current_debt'] + STARTING_BALANCE['long_term_debt']

    for i in range(len(years)):
        current_debt = balance_sheet.loc[i, 'Current_Debt'] + balance_sheet.loc[i, 'Long_Term_Debt']
        debt_payments.append(-(prev_debt - current_debt))
        prev_debt = current_debt

    data['Debt_Payments'] = debt_payments
    data['Financing_Cash_Flow'] = debt_payments

    # Net Cash Flow
    data['Net_Cash_Flow'] = [
        data['Operating_Cash_Flow'][i] + data['Investing_Cash_Flow'][i] + data['Financing_Cash_Flow'][i]
        for i in range(len(years))
    ]

    # Ending Cash
    cash = STARTING_BALANCE['cash']
    ending_cash = []
    for i in range(len(years)):
        cash = cash + data['Net_Cash_Flow'][i]
        ending_cash.append(cash)
    data['Ending_Cash'] = ending_cash

    # Free Cash Flow
    data['Free_Cash_Flow'] = [
        data['Operating_Cash_Flow'][i] + data['CapEx'][i]
        for i in range(len(years))
    ]

    return pd.DataFrame(data)


# ============================================================
# BUILD KEY RATIOS
# ============================================================

def build_ratios(scenario_name, income_stmt, balance_sheet, cash_flow):
    """Calculate key financial ratios."""

    years = PROJECTION_YEARS
    data = {
        'Year': years,
        'Scenario': [scenario_name] * len(years),
    }

    # Profitability Ratios
    data['Gross_Margin'] = income_stmt['Gross_Margin_Pct'].tolist()
    data['EBITDA_Margin'] = income_stmt['EBITDA_Margin_Pct'].tolist()
    data['Net_Margin'] = income_stmt['Net_Margin_Pct'].tolist()

    # Prime Cost (Food + Bev + Labor)
    prime_cost = [
        (income_stmt.loc[i, 'Food_Cost'] + income_stmt.loc[i, 'Beverage_Cost'] + income_stmt.loc[i, 'Labor_Cost']) / income_stmt.loc[i, 'Revenue']
        for i in range(len(years))
    ]
    data['Prime_Cost_Pct'] = prime_cost

    # Liquidity Ratios
    data['Current_Ratio'] = [
        balance_sheet.loc[i, 'Total_Current_Assets'] / balance_sheet.loc[i, 'Total_Current_Liabilities']
        for i in range(len(years))
    ]

    quick_assets = [
        balance_sheet.loc[i, 'Cash'] + balance_sheet.loc[i, 'Accounts_Receivable']
        for i in range(len(years))
    ]
    data['Quick_Ratio'] = [
        quick_assets[i] / balance_sheet.loc[i, 'Total_Current_Liabilities']
        for i in range(len(years))
    ]

    # Leverage Ratios
    data['Debt_to_Equity'] = [
        balance_sheet.loc[i, 'Total_Liabilities'] / balance_sheet.loc[i, 'Total_Equity']
        for i in range(len(years))
    ]

    # Interest Coverage
    data['Interest_Coverage'] = [
        income_stmt.loc[i, 'EBIT'] / max(1, income_stmt.loc[i, 'Interest_Expense'])
        for i in range(len(years))
    ]

    # Return Ratios
    data['ROE'] = [
        income_stmt.loc[i, 'Net_Income'] / balance_sheet.loc[i, 'Total_Equity']
        for i in range(len(years))
    ]

    data['ROA'] = [
        income_stmt.loc[i, 'Net_Income'] / balance_sheet.loc[i, 'Total_Assets']
        for i in range(len(years))
    ]

    return pd.DataFrame(data)


# ============================================================
# BUILD SENSITIVITY ANALYSIS
# ============================================================

def build_sensitivity_analysis(base_revenue):
    """Build sensitivity table for revenue growth and food cost."""

    # Revenue growth sensitivity
    growth_rates = [0.02, 0.04, 0.06, 0.08, 0.10]
    food_costs = [0.29, 0.31, 0.33, 0.35]

    results = []

    for growth in growth_rates:
        for food_cost in food_costs:
            # Quick 5-year projection
            revenue = base_revenue
            total_net_income = 0

            for year in range(5):
                revenue = revenue * (1 + growth)
                cogs = revenue * (food_cost + 0.09)  # Food + bev
                gross_profit = revenue - cogs
                labor = revenue * 0.28
                other_opex = revenue * 0.12
                ebitda = gross_profit - labor - other_opex
                net_income = ebitda * 0.75  # Simplified
                total_net_income += net_income

            results.append({
                'Revenue_Growth': f"{growth:.0%}",
                'Food_Cost_Pct': f"{food_cost:.0%}",
                'Year_5_Revenue': revenue,
                'Total_5Y_Net_Income': total_net_income,
                'Avg_Annual_Net_Income': total_net_income / 5,
            })

    return pd.DataFrame(results)


# ============================================================
# CREATE ASSUMPTIONS SUMMARY
# ============================================================

def create_assumptions_df():
    """Create a summary of assumptions for Excel output."""

    data = []

    for scenario, assumptions in SCENARIOS.items():
        for i, year in enumerate(PROJECTION_YEARS):
            data.append({
                'Scenario': scenario,
                'Year': year,
                'Revenue_Growth': assumptions['revenue_growth'][i],
                'Food_Cost_Pct': assumptions['food_cost_pct'][i],
                'Beverage_Cost_Pct': assumptions['beverage_cost_pct'][i],
                'Labor_Cost_Pct': assumptions['labor_cost_pct'][i],
                'CapEx': assumptions['capex'][i],
            })

    return pd.DataFrame(data)


# ============================================================
# MAIN EXECUTION
# ============================================================

def main():
    """Build complete financial model and export to Excel."""

    print("=" * 60)
    print("FINANCIAL MODEL BUILDER")
    print("=" * 60)

    # Load historical data
    print("\n1. Loading historical data...")
    historical = load_historical_data()
    base_revenue = get_base_revenue()
    print(f"   Base revenue (2023): ${base_revenue:,.0f}")

    # Build projections for each scenario
    all_income_stmts = []
    all_balance_sheets = []
    all_cash_flows = []
    all_ratios = []

    print("\n2. Building projections...")

    for scenario_name, assumptions in SCENARIOS.items():
        print(f"   → {scenario_name} scenario...")

        # Income Statement
        income_stmt = build_income_statement(scenario_name, assumptions, base_revenue)
        all_income_stmts.append(income_stmt)

        # Balance Sheet
        balance_sheet = build_balance_sheet(scenario_name, income_stmt, assumptions)
        all_balance_sheets.append(balance_sheet)

        # Cash Flow
        cash_flow = build_cash_flow(scenario_name, income_stmt, balance_sheet, assumptions)
        all_cash_flows.append(cash_flow)

        # Ratios
        ratios = build_ratios(scenario_name, income_stmt, balance_sheet, cash_flow)
        all_ratios.append(ratios)

    # Combine all scenarios
    income_stmt_all = pd.concat(all_income_stmts, ignore_index=True)
    balance_sheet_all = pd.concat(all_balance_sheets, ignore_index=True)
    cash_flow_all = pd.concat(all_cash_flows, ignore_index=True)
    ratios_all = pd.concat(all_ratios, ignore_index=True)

    # Build sensitivity analysis
    print("\n3. Building sensitivity analysis...")
    sensitivity = build_sensitivity_analysis(base_revenue)

    # Create assumptions summary
    assumptions_df = create_assumptions_df()

    # Build Executive Summary for stakeholders
    print("\n4. Building executive summary...")

    exec_summary_data = []
    for scenario in ['Base', 'Upside', 'Downside']:
        scenario_data = income_stmt_all[income_stmt_all['Scenario'] == scenario]
        year_5 = scenario_data[scenario_data['Year'] == 2028].iloc[0]
        year_1 = scenario_data[scenario_data['Year'] == 2024].iloc[0]

        # Get cash flow data
        cf_data = cash_flow_all[cash_flow_all['Scenario'] == scenario]
        year_5_cf = cf_data[cf_data['Year'] == 2028].iloc[0]

        # Get balance sheet data
        bs_data = balance_sheet_all[balance_sheet_all['Scenario'] == scenario]
        year_5_bs = bs_data[bs_data['Year'] == 2028].iloc[0]

        # 5-year totals
        total_revenue = scenario_data['Revenue'].sum()
        total_net_income = scenario_data['Net_Income'].sum()
        total_fcf = cf_data['Free_Cash_Flow'].sum()

        exec_summary_data.append({
            'Scenario': scenario,
            'Year_5_Revenue': year_5['Revenue'],
            'Year_5_EBITDA': year_5['EBITDA'],
            'Year_5_EBITDA_Margin': year_5['EBITDA_Margin_Pct'],
            'Year_5_Net_Income': year_5['Net_Income'],
            'Year_5_Net_Margin': year_5['Net_Margin_Pct'],
            'Year_5_Free_Cash_Flow': year_5_cf['Free_Cash_Flow'],
            'Year_5_Ending_Cash': year_5_cf['Ending_Cash'],
            '5Y_Total_Revenue': total_revenue,
            '5Y_Total_Net_Income': total_net_income,
            '5Y_Total_Free_Cash_Flow': total_fcf,
            'Revenue_CAGR': ((year_5['Revenue'] / year_1['Revenue']) ** 0.25 - 1),
        })

    exec_summary = pd.DataFrame(exec_summary_data)

    # Export to Excel
    print("\n5. Exporting to Excel...")

    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        from openpyxl.styles import numbers
        # EXECUTIVE SUMMARY FIRST - This is what stakeholders need
        exec_summary.to_excel(writer, sheet_name='Executive Summary', index=False)

        # Scenario Comparison (detailed pivot view)
        comparison_revenue = income_stmt_all.pivot(index='Year', columns='Scenario', values='Revenue')
        comparison_ebitda = income_stmt_all.pivot(index='Year', columns='Scenario', values='EBITDA')
        comparison_net = income_stmt_all.pivot(index='Year', columns='Scenario', values='Net_Income')
        comparison_fcf = cash_flow_all.pivot(index='Year', columns='Scenario', values='Free_Cash_Flow')

        # Combine into one comparison sheet
        comparison_revenue.columns = [f'{col}_Revenue' for col in comparison_revenue.columns]
        comparison_ebitda.columns = [f'{col}_EBITDA' for col in comparison_ebitda.columns]
        comparison_net.columns = [f'{col}_Net_Income' for col in comparison_net.columns]
        comparison_fcf.columns = [f'{col}_FCF' for col in comparison_fcf.columns]

        full_comparison = pd.concat([comparison_revenue, comparison_ebitda, comparison_net, comparison_fcf], axis=1)
        full_comparison.to_excel(writer, sheet_name='Scenario Comparison')

        # Assumptions
        assumptions_df.to_excel(writer, sheet_name='Assumptions', index=False)

        # Historical
        historical.to_excel(writer, sheet_name='Historical', index=False)

        # Income Statement
        income_stmt_all.to_excel(writer, sheet_name='Income Statement', index=False)

        # Balance Sheet
        balance_sheet_all.to_excel(writer, sheet_name='Balance Sheet', index=False)

        # Cash Flow
        cash_flow_all.to_excel(writer, sheet_name='Cash Flow', index=False)

        # Ratios
        ratios_all.to_excel(writer, sheet_name='Key Ratios', index=False)

        # Sensitivity
        sensitivity.to_excel(writer, sheet_name='Sensitivity', index=False)

        # Apply formatting to all sheets
        print("\n6. Applying formatting...")

        # Currency columns (no decimals)
        currency_cols = [
            'Revenue', 'Food_Cost', 'Beverage_Cost', 'COGS', 'Gross_Profit',
            'Labor_Cost', 'Rent', 'Utilities', 'Marketing', 'Insurance',
            'Supplies', 'Repairs_Maintenance', 'Credit_Card_Fees', 'Other_OpEx',
            'Total_OpEx', 'EBITDA', 'Depreciation', 'EBIT', 'Interest_Expense',
            'EBT', 'Income_Tax', 'Net_Income', 'Cash', 'Accounts_Receivable',
            'Inventory', 'Total_Current_Assets', 'Fixed_Assets_Net', 'Total_Assets',
            'Accounts_Payable', 'Current_Debt', 'Total_Current_Liabilities',
            'Long_Term_Debt', 'Total_Liabilities', 'Retained_Earnings',
            'Total_Equity', 'Total_Liab_Equity', 'Operating_Cash_Flow',
            'CapEx', 'Investing_Cash_Flow', 'Debt_Payments', 'Financing_Cash_Flow',
            'Net_Cash_Flow', 'Ending_Cash', 'Free_Cash_Flow',
            'Year_5_Revenue', 'Year_5_EBITDA', 'Year_5_Net_Income',
            'Year_5_Free_Cash_Flow', 'Year_5_Ending_Cash',
            '5Y_Total_Revenue', '5Y_Total_Net_Income', '5Y_Total_Free_Cash_Flow',
            'Year_5_Revenue', 'Total_5Y_Net_Income', 'Avg_Annual_Net_Income',
        ]

        # Percentage columns
        pct_cols = [
            'Gross_Margin_Pct', 'EBITDA_Margin_Pct', 'Net_Margin_Pct',
            'Year_5_EBITDA_Margin', 'Year_5_Net_Margin', 'Revenue_CAGR',
            'Revenue_Growth', 'Food_Cost_Pct', 'Beverage_Cost_Pct', 'Labor_Cost_Pct',
            'Gross_Margin', 'EBITDA_Margin', 'Net_Margin', 'Prime_Cost_Pct',
            'ROE', 'ROA', 'Discount',
        ]

        # Format each sheet
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]

            # Get header row to find column indices
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value] = col_idx

            # Apply currency format
            for col_name in currency_cols:
                if col_name in headers:
                    col_idx = headers[col_name]
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        if cell.value is not None:
                            cell.number_format = '"$"#,##0'

            # Apply percentage format
            for col_name in pct_cols:
                if col_name in headers:
                    col_idx = headers[col_name]
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        if cell.value is not None:
                            cell.number_format = '0.0%'

            # Also check for columns that contain currency/percentage keywords
            for header, col_idx in headers.items():
                if header and ('Revenue' in str(header) or 'EBITDA' in str(header) or
                              'Income' in str(header) or 'FCF' in str(header) or
                              'Cash' in str(header)):
                    if 'Margin' not in str(header) and 'Pct' not in str(header) and 'CAGR' not in str(header):
                        for row in range(2, ws.max_row + 1):
                            cell = ws.cell(row=row, column=col_idx)
                            if cell.value is not None and isinstance(cell.value, (int, float)):
                                cell.number_format = '"$"#,##0'

    print(f"\n   Saved: {OUTPUT_FILE}")

    # Print summary
    print("\n" + "=" * 60)
    print("MODEL SUMMARY")
    print("=" * 60)

    for scenario in ['Base', 'Upside', 'Downside']:
        scenario_data = income_stmt_all[income_stmt_all['Scenario'] == scenario]
        year_5 = scenario_data[scenario_data['Year'] == 2028].iloc[0]

        print(f"\n{scenario} Case (2028):")
        print(f"  Revenue:     ${year_5['Revenue']:>12,.0f}")
        print(f"  EBITDA:      ${year_5['EBITDA']:>12,.0f}  ({year_5['EBITDA_Margin_Pct']:.1%})")
        print(f"  Net Income:  ${year_5['Net_Income']:>12,.0f}  ({year_5['Net_Margin_Pct']:.1%})")

    print("\n" + "=" * 60)
    print("COMPLETE")
    print("=" * 60)


if __name__ == '__main__':
    main()
